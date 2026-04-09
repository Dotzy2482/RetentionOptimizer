"""Export utilities: Excel customer report and PDF dashboard summary."""

import os
from datetime import datetime

import pandas as pd
from sqlalchemy import text
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from data.database import engine


def export_customers_excel(file_path: str) -> int:
    """Export all customers with RFM, segment, loyalty, churn to Excel.
    Returns number of rows exported.
    """
    with engine.connect() as conn:
        df = pd.read_sql(text("""
            SELECT
                c.customer_id   AS "Musteri ID",
                c.country       AS "Ulke",
                r.recency       AS "Recency",
                r.frequency     AS "Frequency",
                r.monetary      AS "Monetary",
                r.r_score       AS "R Score",
                r.f_score       AS "F Score",
                r.m_score       AS "M Score",
                r.loyalty_score AS "Loyalty Score",
                s.segment_label AS "Segment",
                s.churn_probability AS "Churn Olasiligi"
            FROM customers c
            LEFT JOIN rfm_scores r  ON c.customer_id = r.customer_id
            LEFT JOIN segments s    ON c.customer_id = s.customer_id
            ORDER BY r.loyalty_score DESC
        """), conn)

    if df.empty:
        raise ValueError("Disa aktarilacak veri bulunamadi.")

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Musteriler")

        ws = writer.sheets["Musteriler"]

        # Style header row
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            bottom=Side(style="thin", color="E5E5EA"),
            right=Side(style="thin", color="E5E5EA"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 25)

        # Format churn probability as percentage
        churn_col_idx = df.columns.get_loc("Churn Olasiligi") + 1
        for row in ws.iter_rows(min_row=2, min_col=churn_col_idx, max_col=churn_col_idx):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = "0.00%"

    return len(df)


def export_dashboard_pdf(file_path: str) -> None:
    """Export dashboard summary with metrics and charts to PDF."""
    from matplotlib.backends.backend_pdf import PdfPages
    from matplotlib.figure import Figure

    with engine.connect() as conn:
        rfm_df = pd.read_sql(text("SELECT * FROM rfm_scores"), conn)
        seg_df = pd.read_sql(text("""
            SELECT s.*, r.loyalty_score, r.recency, r.frequency, r.monetary
            FROM segments s
            JOIN rfm_scores r ON s.customer_id = r.customer_id
        """), conn)

    if rfm_df.empty:
        raise ValueError("Disa aktarilacak veri bulunamadi.")

    with PdfPages(file_path) as pdf:
        # Page 1: Summary metrics + charts
        fig = Figure(figsize=(8.27, 11.69))  # A4
        fig.set_facecolor("white")

        # Title
        fig.text(0.5, 0.94, "Retention Optimization System", ha="center",
                 fontsize=20, fontweight="bold", color="#7C3AED")
        fig.text(0.5, 0.91, f"Dashboard Raporu - {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                 ha="center", fontsize=11, color="#8E8E93")

        # Summary text
        total = len(rfm_df)
        avg_loyalty = rfm_df["loyalty_score"].mean()
        lines = [
            f"Toplam Musteri: {total:,}",
            f"Ortalama Loyalty Score: {avg_loyalty:.1f}",
        ]
        if not seg_df.empty:
            seg_counts = seg_df["segment_label"].value_counts()
            for label, count in seg_counts.items():
                pct = count / total * 100
                lines.append(f"  {label}: {count:,} (%{pct:.1f})")

            avg_churn = seg_df["churn_probability"].mean()
            if pd.notna(avg_churn):
                lines.append(f"Ortalama Churn Olasiligi: %{avg_churn * 100:.1f}")

        y = 0.85
        for line in lines:
            fig.text(0.1, y, line, fontsize=12, color="#1C1C1E")
            y -= 0.03

        # Loyalty histogram
        ax1 = fig.add_axes([0.1, 0.45, 0.8, 0.25])
        ax1.hist(rfm_df["loyalty_score"].values, bins=20, color="#A78BFA",
                 edgecolor="#7C3AED", alpha=0.8)
        ax1.set_title("Loyalty Score Dagilimi", fontsize=11, fontweight="bold", color="#1C1C1E")
        ax1.set_xlabel("Loyalty Score", color="#8E8E93")
        ax1.set_ylabel("Musteri Sayisi", color="#8E8E93")
        ax1.grid(axis="y", alpha=0.2)

        # Segment pie
        if not seg_df.empty:
            ax2 = fig.add_axes([0.15, 0.08, 0.7, 0.30])
            counts = seg_df["segment_label"].value_counts()
            colors = ["#FF3B30", "#FF9500", "#A78BFA", "#34C759"]
            ax2.pie(counts.values, labels=counts.index, autopct="%1.1f%%",
                    colors=colors[:len(counts)], startangle=90,
                    textprops={"fontsize": 9},
                    wedgeprops={"edgecolor": "white", "linewidth": 2})
            ax2.set_title("Segment Dagilimi", fontsize=11, fontweight="bold", color="#1C1C1E")

        pdf.savefig(fig)
